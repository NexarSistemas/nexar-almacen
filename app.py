from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session, send_file, send_from_directory
from datetime import date, datetime, timedelta
from functools import wraps
import json
import os
import sys
import signal
import shutil
import glob
import threading
import secrets

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import database as db

# Versión dinámica: se lee del archivo VERSION (fuente única de verdad)
def _read_version():
    try:
        v = open(os.path.join(os.path.dirname(__file__), 'VERSION')).read().strip()
        return v if v else "1.7.0"
    except Exception: # Fallback si no existe el archivo VERSION
        return "1.7.0"

APP_VERSION = _read_version()

app = Flask(__name__)

def get_secret_key():
    key = os.getenv("SECRET_KEY")
    if key:
        return key

    # carpeta del usuario
    config_dir = os.path.join(os.path.expanduser("~"), ".nexar")
    os.makedirs(config_dir, exist_ok=True)

    config_path = os.path.join(config_dir, "config.json")

    if os.path.exists(config_path):
        with open(config_path, "r") as f:
            data = json.load(f)
            if "SECRET_KEY" in data:
                return data["SECRET_KEY"]

    key = secrets.token_hex(32)

    with open(config_path, "w") as f:
        json.dump({"SECRET_KEY": key}, f)

    return key


SECRET_KEY = os.getenv("SECRET_KEY", "").strip()

if not SECRET_KEY:
    raise RuntimeError("SECRET_KEY no definida. Configurar variable de entorno.")

app.config["SECRET_KEY"] = SECRET_KEY

def get_public_key():
    import os

    # 1. entorno
    key = os.getenv("PUBLIC_KEY")
    if key:
        return key

    # 2. archivo
    path = "keys/public_key.asc"
    if os.path.exists(path):
        with open(path, "r") as f:
            return f.read()

    # 3. error
    raise RuntimeError("Clave pública no encontrada")

app.jinja_env.add_extension('jinja2.ext.loopcontrols')


@app.route('/favicon.ico')
def favicon():
    return send_from_directory(
        os.path.join(app.root_path, 'static/icons'),
        'nexar_almacen.ico',
        mimetype='image/vnd.microsoft.icon'
    )

# ─── BACKUP SCHEDULER ────────────────────────────────────────────────────────
_backup_timer = None

def get_backup_dir():
    cfg = db.get_config()
    custom = cfg.get('backup_dir', '').strip()
    if custom and os.path.isabs(custom):
        return custom
    base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, 'respaldo')

def hacer_backup(manual=False):
    """Copies almacen.db → respaldo/almacen_YYYY-MM-DD_HH-MM.db"""
    try:
        db_path = db.DB_PATH if hasattr(db, 'DB_PATH') else os.path.join(
            os.path.dirname(os.path.abspath(db.__file__)), 'almacen.db')
        # Try to get DB_PATH from the module source if not exposed
        if not os.path.exists(db_path):
            # Fallback: same dir as database.py
            db_path = os.path.join(os.path.dirname(os.path.abspath(db.__file__)), 'almacen.db')
        if not os.path.exists(db_path):
            return False, 'No se encontró la base de datos.'
        backup_dir = get_backup_dir()
        os.makedirs(backup_dir, exist_ok=True)
        ts = datetime.now().strftime('%Y-%m-%d_%H-%M')
        dest = os.path.join(backup_dir, f'almacen_{ts}.db')
        shutil.copy2(db_path, dest)
        # Keep only last N backups
        cfg = db.get_config()
        keep = int(cfg.get('backup_keep', '10'))
        all_bkp = sorted(glob.glob(os.path.join(backup_dir, 'almacen_*.db')))
        for old in all_bkp[:-keep]:
            try:
                os.remove(old)
            except Exception:
                pass
        db.set_config({'backup_ultimo': datetime.now().strftime('%Y-%m-%d %H:%M:%S')})
        return True, dest
    except Exception as e:
        return False, str(e)

def _schedule_backup():
    global _backup_timer
    try:
        cfg = db.get_config()
        interval_h = float(cfg.get('backup_intervalo_h', '24'))
        if interval_h <= 0:
            return
        hacer_backup()
        interval_s = interval_h * 3600
        _backup_timer = threading.Timer(interval_s, _schedule_backup)
        _backup_timer.daemon = True
        _backup_timer.start()
    except Exception:
        pass

def iniciar_backup_scheduler():
    global _backup_timer
    if _backup_timer:
        _backup_timer.cancel()
    _backup_timer = threading.Timer(5, _schedule_backup)   # first run after 5s
    _backup_timer.daemon = True
    _backup_timer.start()

@app.template_filter('enumerate')
def jinja_enumerate(iterable):
    return enumerate(iterable)

# ─── AUTH DECORATORS ──────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user' not in session:
            flash('⚠ Debés iniciar sesión para acceder.', 'warning')
            return redirect(url_for('login', next=request.path))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user' not in session:
            return redirect(url_for('login'))
        if session['user'].get('rol') != 'admin':
            flash('⛔ Acceso restringido: se requieren permisos de Administrador.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

def demo_venta_check(f):
    """Block new sales when demo has expired (30-day time limit)."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if db.is_demo_mode():
            status = db.get_demo_status()
            if status.get('vencido'):
                return jsonify({'ok': False, 'demo_blocked': True,
                    'msg': 'Tu período de prueba de 30 días ha vencido. Activá tu licencia en Sistema → Licencia.',
                    'redirect': '/licencia'}), 403
        return f(*args, **kwargs)
    return decorated

_scheduler_started = False

@app.before_request
def before():
    global _scheduler_started
    db.init_db()
    if not _scheduler_started:
        _scheduler_started = True
        iniciar_backup_scheduler()
    # Allow public routes without login
    public = {'/login', '/static'}
    if any(request.path.startswith(p) for p in public):
        return
    # Invalidar sesión si el sistema fue apagado desde otra sesión
    if 'user' in session:
        try:
            inv_at = db.get_config().get('sessions_invalidated_at', '')
            login_at = session.get('login_date', '')
            if inv_at and login_at and login_at < inv_at:
                session.clear()
        except Exception:
            pass

    if 'user' not in session and request.endpoint not in ('login', 'apagar_rapido', 'static', None):
        return redirect(url_for('login', next=request.path))

@app.context_processor
def inject_globals():
    if 'user' not in session:
        return {}
    try:
        alertas = db.get_alertas_count()
        fact_venc = db.q("""SELECT COUNT(*) as n FROM facturas_proveedores
            WHERE fecha_vencimiento <= date('now','+30 days') AND importe > pagado""", fetchone=True)
        cfg = db.get_config()
        demo_status = db.get_demo_status()

        return {
            'alertas_sidebar': {
                'sin_stock': alertas['sin_stock'],
                'critico': alertas['critico'],
                'facturas_por_vencer': fact_venc['n'] if fact_venc else 0,
            },
            'cfg_nombre': cfg.get('nombre_negocio', 'Mi Almacén'),
            'current_user': session.get('user', {}),
            'is_admin': session.get('user', {}).get('rol') == 'admin',
            'demo_status': demo_status,
            'app_version': APP_VERSION,

            # 🔥 LICENCIAS (correcto)
            'tier': db.get_tier(),           # 'DEMO' | 'BASICA' | 'PRO'
            'pro_expired': db.is_pro_expired(),

        }
    except Exception as e:
        print("ERROR inject_globals:", e)  # 👈 clave para debug
        return {}

# ─── HELPERS ─────────────────────────────────────────────────────────────────
def fmt_ars(v):
    try:
        return f"${float(v):,.2f}".replace(',','X').replace('.',',').replace('X','.')
    except:
        return "$0,00"

app.jinja_env.globals['fmt_ars'] = fmt_ars
app.jinja_env.globals['today'] = lambda: date.today().isoformat()
app.jinja_env.globals['now'] = lambda: datetime.now().strftime('%H:%M')

# ─── LOGIN / LOGOUT ──────────────────────────────────────────────────────────
@app.route('/login', methods=['GET','POST'])
def login():
    if 'user' in session:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username','').strip()
        password = request.form.get('password','')
        user = db.verificar_password(username, password)
        if user:
            from datetime import datetime as _dt
            session['user'] = {
                'id':       user['id'],
                'username': user['username'],
                'nombre':   user['nombre_completo'] or user['username'],
                'rol':      user['rol'],
            }
            session['login_date'] = _dt.now().isoformat()
            next_url = request.form.get('next') or url_for('dashboard')
            flash(f'✅ Bienvenido, {user["nombre_completo"] or username}!', 'success')
            return redirect(next_url)
        flash('❌ Usuario o contraseña incorrectos.', 'danger')
    next_url = request.args.get('next', '')
    return render_template('login.html', next=next_url, app_version=APP_VERSION)

@app.route('/logout')
def logout():
    nombre = session.get('user', {}).get('nombre', '')
    session.clear()
    flash(f'👋 Sesión cerrada. ¡Hasta pronto, {nombre}!', 'info')
    return redirect(url_for('login'))

# ─── LICENCIA / ACTIVACIÓN ───────────────────────────────────────────────────
@app.route('/licencia')
@login_required
def licencia():
    status     = db.get_demo_status()
    machine_id = db.get_machine_id()
    lic_info   = db.get_license_info() if not status['demo'] else None

    # Calcular uso actual para mostrar barras de progreso en plan Básico
    tier_limits = {}
    if lic_info and lic_info.get('tier') == 'BASICA':
        r_off  = db.q(
            "SELECT COUNT(*) FROM productos WHERE activo=1 "
            "AND codigo_barras != '' "
            "AND (codigo_barras LIKE '779%' OR codigo_barras LIKE '780%')",
            fetchone=True
        )
        r_cli  = db.q("SELECT COUNT(*) FROM clientes WHERE activo=1", fetchone=True)
        r_prov = db.q("SELECT COUNT(*) FROM proveedores WHERE activo=1", fetchone=True)
        tier_limits = {
            'productos_off_actual': r_off[0]  if r_off  else 0,
            'clientes_actual':      r_cli[0]  if r_cli  else 0,
            'proveedores_actual':   r_prov[0] if r_prov else 0,
        }

    return render_template('licencia.html',
                           status=status,
                           machine_id=machine_id,
                           lic_info=lic_info,
                           tier_limits=tier_limits)

@app.route('/licencia/activar', methods=['POST'])
@admin_required
def licencia_activar():
    token = request.form.get('token', '').strip()
    if not token:
        flash('❌ No ingresaste ningún token. Pegá el código que recibiste por WhatsApp.', 'danger')
        return redirect(url_for('licencia'))

    # Validar el token sin activar todavía para leer el tier
    ok, msg, data = db.validar_licencia_rsa(token)
    if not ok:
        flash(f'❌ {msg}', 'danger')
        return redirect(url_for('licencia'))

    tier_token = data.get('tier', 'BASICA')

    # Si el token es PRO, verificar que ya tenga Básica activa
    if tier_token == 'PRO':
        cfg = db.get_config()
        tiene_basica = (
            cfg.get('demo_mode', '1') == '0' and
            cfg.get('license_tier', '') == 'BASICA'
        )
        if not tiene_basica:
            flash(
                '⚠ Para activar el Plan Pro primero debés activar el Plan Básico. '
                'Contactá al desarrollador para adquirir el bundle Básica + Pro.',
                'warning'
            )
            return redirect(url_for('licencia'))

    # Activar la licencia
    ok, msg = db.activar_licencia(token)
    if ok:
        tier = db.get_tier()
        if tier == 'PRO':
            flash('🎉 ¡Plan Pro activado! Acceso ilimitado + actualizaciones habilitadas.', 'success')
        else:
            flash('✅ ¡Plan Básico activado! Sistema habilitado con funciones del plan Básico.', 'success')
    else:
        flash(f'❌ {msg}', 'danger')
    return redirect(url_for('licencia'))

# ─── USUARIOS (admin only) ───────────────────────────────────────────────────
@app.route('/usuarios')
@admin_required
def usuarios():
    users = db.get_usuarios()
    return render_template('usuarios.html', users=users)

@app.route('/usuarios/nuevo', methods=['POST'])
@admin_required
def usuario_nuevo():
    ok = db.crear_usuario(
        request.form.get('username',''),
        request.form.get('password',''),
        request.form.get('rol','usuario'),
        request.form.get('nombre_completo','')
    )
    if ok:
        flash('✅ Usuario creado correctamente.', 'success')
    else:
        flash('❌ El nombre de usuario ya existe.', 'danger')
    return redirect(url_for('usuarios'))

@app.route('/usuarios/<int:uid>/editar', methods=['POST'])
@admin_required
def usuario_editar(uid):
    db.editar_usuario(uid, request.form.get('nombre_completo',''), request.form.get('rol','usuario'))
    flash('✅ Usuario actualizado.', 'success')
    return redirect(url_for('usuarios'))

@app.route('/usuarios/<int:uid>/password', methods=['POST'])
@admin_required
def usuario_password(uid):
    nueva = request.form.get('password','')
    if len(nueva) < 4:
        flash('❌ La contraseña debe tener al menos 4 caracteres.', 'danger')
    else:
        db.cambiar_password(uid, nueva)
        flash('✅ Contraseña cambiada.', 'success')
    return redirect(url_for('usuarios'))

@app.route('/usuarios/<int:uid>/toggle', methods=['POST'])
@admin_required
def usuario_toggle(uid):
    if uid == session['user']['id']:
        flash('⚠ No podés desactivar tu propia cuenta.', 'warning')
    else:
        db.toggle_usuario(uid)
        flash('✅ Estado del usuario actualizado.', 'success')
    return redirect(url_for('usuarios'))

@app.route('/usuarios/<int:uid>/eliminar', methods=['POST'])
@admin_required
def usuario_eliminar(uid):
    if uid == session['user']['id']:
        flash('⚠ No podés eliminar tu propia cuenta.', 'warning')
    else:
        db.delete_usuario(uid)
        flash('🗑 Usuario eliminado.', 'warning')
    return redirect(url_for('usuarios'))

# ─── AYUDA / ACERCA ──────────────────────────────────────────────────────────
@app.route('/ayuda')
@login_required
def ayuda():
    return render_template('ayuda.html')

@app.route('/acerca')
@login_required
def acerca():
    entries = db.get_changelog()
    return render_template('acerca.html', app_version=APP_VERSION, changelog=entries)

# ─── DASHBOARD ───────────────────────────────────────────────────────────────
@app.route('/')
@login_required
def dashboard():
    kpis = db.get_dashboard_kpis()
    ventas_mes_data = db.get_ventas_por_mes()
    meses_labels = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
    ventas_chart = [ventas_mes_data.get(m, {}).get('total', 0) for m in range(1, 13)]
    alertas_stock = db.get_stock_full(alerta_only=True)
    facturas_venc = db.q("""SELECT fp.*,p.nombre as prov FROM facturas_proveedores fp
        JOIN proveedores p ON p.id=fp.proveedor_id
        WHERE fp.fecha_vencimiento<=date('now','+30 days') AND fp.importe>fp.pagado
        ORDER BY fp.fecha_vencimiento LIMIT 5""")
    return render_template('dashboard.html',
        kpis=kpis, ventas_chart=json.dumps(ventas_chart),
        meses_labels=json.dumps(meses_labels),
        alertas_stock=alertas_stock[:10],
        facturas_venc=facturas_venc)

# ─── PRODUCTOS ───────────────────────────────────────────────────────────────
@app.route('/productos')
@login_required
def productos():
    search = request.args.get('q', '')
    prods = db.get_productos(search=search)
    cats = db.get_categorias()
    return render_template('productos.html', productos=prods, categorias=cats, search=search)

@app.route('/productos/nuevo', methods=['GET','POST'])
@login_required
def producto_nuevo():
    cats = db.get_categorias()
    # Demo check (v1.3: time-based)
    if request.method == 'POST':
        if db.is_demo_mode():
            status = db.get_demo_status()
            if status.get('vencido'):
                flash('⚠ Tu período de prueba de 30 días ha vencido. Activá tu licencia para continuar.', 'warning')
                return redirect(url_for('licencia'))
        data = request.form.to_dict()
        db.add_producto(data)
        flash('✅ Producto creado correctamente', 'success')
        return redirect(url_for('productos'))
    return render_template('producto_form.html', producto=None, categorias=cats, accion='Nuevo')

@app.route('/productos/<int:pid>/editar', methods=['GET','POST'])
@login_required
def producto_editar(pid):
    prod = db.get_producto(pid)
    stock_row = db.q("SELECT * FROM stock WHERE producto_id=?", (pid,), fetchone=True)
    cats = db.get_categorias()
    if request.method == 'POST':
        data = request.form.to_dict()
        db.update_producto(pid, data)
        if stock_row:
            db.update_stock_item(pid,
                stock_actual=float(data.get('stock_actual', stock_row['stock_actual'])),
                stock_minimo=float(data.get('stock_minimo', stock_row['stock_minimo'])),
                stock_maximo=float(data.get('stock_maximo', stock_row['stock_maximo'])))
        flash('✅ Producto actualizado', 'success')
        return redirect(url_for('productos'))
    return render_template('producto_form.html', producto=prod, stock=stock_row, categorias=cats, accion='Editar')

@app.route('/productos/<int:pid>/eliminar', methods=['POST'])
@login_required
def producto_eliminar(pid):
    db.delete_producto(pid)
    flash('Producto desactivado', 'info')
    return redirect(url_for('productos'))

@app.route('/api/producto/buscar')
@login_required
def api_producto_buscar():
    codigo = request.args.get('codigo', '').strip()
    if not codigo:
        return jsonify({'ok': False, 'msg': 'Código vacío'})
    p = db.get_producto_by_codigo(codigo)
    if not p:
        return jsonify({'ok': False, 'msg': f'Producto no encontrado: {codigo}'})
    s = db.q("SELECT stock_actual FROM stock WHERE producto_id=?", (p['id'],), fetchone=True)
    return jsonify({
        'ok': True,
        'id': p['id'],
        'codigo_interno': p['codigo_interno'],
        'descripcion': p['descripcion'],
        'categoria': p['categoria'],
        'unidad': p['unidad'],
        'precio_venta': p['precio_venta'],
        'por_peso': p['por_peso'],
        'stock_actual': s['stock_actual'] if s else 0,
    })

@app.route('/api/categorias/nueva', methods=['POST'])
@login_required
def api_categoria_nueva():
    nombre = request.json.get('nombre','').strip()
    if nombre:
        db.add_categoria(nombre)
        return jsonify({'ok': True})
    return jsonify({'ok': False})

# ─── STOCK ───────────────────────────────────────────────────────────────────
@app.route('/stock')
@login_required
def stock():
    search = request.args.get('q', '')
    alerta = request.args.get('alerta', '')
    rows = db.get_stock_full(search=search, alerta_only=bool(alerta))
    alertas = db.get_alertas_count()
    proveedores = db.get_proveedores(activo_only=True)
    return render_template('stock.html', rows=rows, alertas=alertas,
                           search=search, alerta=alerta, proveedores=proveedores)

@app.route('/stock/<int:pid>/editar', methods=['POST'])
@login_required
def stock_editar(pid):
    data = request.form.to_dict()
    db.update_stock_item(pid,
        stock_actual=float(data.get('stock_actual',0)),
        stock_minimo=float(data.get('stock_minimo',5)),
        stock_maximo=float(data.get('stock_maximo',50)),
        proveedor=data.get('proveedor_habitual',''))
    flash('✅ Stock actualizado', 'success')
    return redirect(url_for('stock'))

# ─── PUNTO DE VENTA ──────────────────────────────────────────────────────────
@app.route('/venta')
@app.route('/punto_venta')
@login_required
def punto_venta():
    clientes = db.get_clientes()
    medios = ['Efectivo','Débito','Crédito','Transferencia','QR / Billetera Virtual','Cuenta Corriente']
    return render_template('punto_venta.html', clientes=clientes, medios=medios)

@app.route('/venta/procesar', methods=['POST'])
@login_required
def venta_procesar():
    # Demo mode check (v1.3: time-based)
    if db.is_demo_mode():
        status = db.get_demo_status()
        if status.get('vencido'):
            return jsonify({'ok': False, 'demo_blocked': True,
                'msg': 'Tu período de prueba de 30 días ha vencido. Activá tu licencia en Sistema → Licencia.',
                'redirect': '/licencia'})
    data = request.json
    items = data.get('items', [])
    if not items:
        return jsonify({'ok': False, 'msg': 'No hay productos en el ticket'})
    try:
        vid, ticket = db.crear_venta(
            items=items,
            cliente_nombre=data.get('cliente_nombre', 'Mostrador'),
            medio_pago=data.get('medio_pago', 'Efectivo'),
            descuento_adicional=float(data.get('descuento_adicional', 0)),
            vendedor=data.get('vendedor', ''),
            cliente_id=int(data.get('cliente_id', 0)),
        )
        return jsonify({'ok': True, 'vid': vid, 'ticket': ticket})
    except Exception as e:
        return jsonify({'ok': False, 'msg': str(e)})

@app.route('/venta/<int:vid>/ticket')
@login_required
def venta_ticket(vid):
    v = db.q("SELECT * FROM ventas WHERE id=?", (vid,), fetchone=True)
    detalle = db.get_venta_detalle(vid)
    cfg = db.get_config()
    return render_template('ticket.html', venta=v, detalle=detalle, cfg=cfg)

# ─── HISTORIAL VENTAS ─────────────────────────────────────────────────────────
@app.route('/historial')
@login_required
def historial():
    search = request.args.get('q', '')
    fecha_desde = request.args.get('desde', '')
    fecha_hasta = request.args.get('hasta', '')
    ventas = db.get_ventas(search=search, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)
    total_filtro = sum(v['total'] for v in ventas)
    return render_template('historial.html', ventas=ventas, search=search,
                           fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
                           total_filtro=total_filtro)

@app.route('/historial/<int:vid>')
@login_required
def historial_detalle(vid):
    v = db.q("SELECT * FROM ventas WHERE id=?", (vid,), fetchone=True)
    detalle = db.get_venta_detalle(vid)
    return render_template('ticket.html', venta=v, detalle=detalle, cfg=db.get_config())

# ─── COMPRAS ─────────────────────────────────────────────────────────────────
@app.route('/compras')
@login_required
def compras():
    search = request.args.get('q', '')
    rows = db.get_compras(search=search)
    provs = db.get_proveedores()
    prods = db.get_productos()
    return render_template('compras.html', compras=rows, proveedores=provs, productos=prods, search=search)

@app.route('/compras/nueva', methods=['POST'])
@login_required
def compra_nueva():
    data = request.form.to_dict()
    # Get product info
    pid = int(data.get('producto_id', 0))
    if pid:
        p = db.get_producto(pid)
        if p:
            data['codigo_interno'] = p['codigo_interno']
            data['descripcion'] = p['descripcion']
    prov_id = int(data.get('proveedor_id', 0))
    if prov_id:
        prov = db.get_proveedor(prov_id)
        if prov:
            data['proveedor_nombre'] = prov['nombre']
    db.registrar_compra(data)
    flash('✅ Compra registrada y stock actualizado', 'success')
    return redirect(url_for('compras'))

# ─── CAJA ────────────────────────────────────────────────────────────────────
@app.route('/caja')
@login_required
def caja():
    caja_hoy = db.get_caja_hoy()
    historial = db.get_caja_historial(30)
    cats_gasto = ['Alquiler','Servicios','Internet/Teléfono','Sueldos','Impuestos/Tasas',
                  'Mantenimiento','Transporte/Flete','Publicidad','Seguros','Descartables/Insumos','Otros']
    medios = ['Efectivo','Débito','Crédito','Transferencia','QR / Billetera Virtual']
    return render_template('caja.html', caja=caja_hoy, historial=historial,
                           cats_gasto=cats_gasto, medios=medios)

@app.route('/caja/abrir', methods=['POST'])
@login_required
def caja_abrir():
    data = request.form
    db.abrir_caja(float(data.get('saldo_apertura',0)), data.get('responsable',''))
    flash('✅ Caja abierta', 'success')
    return redirect(url_for('caja'))

@app.route('/caja/cerrar', methods=['POST'])
@login_required
def caja_cerrar():
    data = request.form
    db.cerrar_caja(float(data.get('saldo_real',0)), data.get('responsable',''))
    flash('✅ Caja cerrada correctamente', 'success')
    return redirect(url_for('caja'))

@app.route('/caja/gasto', methods=['POST'])
@login_required
def caja_gasto():
    data = request.form.to_dict()
    data['tipo'] = 'Gasto'
    db.add_gasto(data)
    flash('✅ Gasto registrado', 'success')
    return redirect(url_for('caja'))

# ─── GASTOS ──────────────────────────────────────────────────────────────────
@app.route('/gastos')
@login_required
def gastos():
    search = request.args.get('q', '')
    rows = db.get_gastos(search=search)
    cats = ['Alquiler','Servicios','Internet/Teléfono','Sueldos','Impuestos/Tasas',
            'Mantenimiento','Transporte/Flete','Publicidad','Seguros','Descartables/Insumos','Otros']
    medios = ['Efectivo','Débito','Crédito','Transferencia','QR / Billetera Virtual']
    # Summary
    total_gasto = sum(r['monto'] for r in rows if r['tipo']=='Gasto')
    total_prescind = sum(r['monto'] for r in rows if r['tipo']=='Gasto' and r['necesario']=='NO (prescindible)')
    return render_template('gastos.html', gastos=rows, search=search, cats=cats, medios=medios,
                           total_gasto=total_gasto, total_prescind=total_prescind)

@app.route('/gastos/nuevo', methods=['POST'])
@login_required
def gasto_nuevo():
    db.add_gasto(request.form.to_dict())
    flash('✅ Gasto registrado', 'success')
    return redirect(url_for('gastos'))

@app.route('/gastos/<int:gid>/eliminar', methods=['POST'])
@login_required
def gasto_eliminar(gid):
    db.delete_gasto(gid)
    flash('Gasto eliminado', 'info')
    return redirect(url_for('gastos'))

@app.route('/gastos/<int:gid>/editar', methods=['POST'])
@login_required
def gasto_editar(gid):
    db.update_gasto(gid, request.form.to_dict())
    flash('✅ Gasto actualizado', 'success')
    return redirect(url_for('gastos'))

# ─── CC CLIENTES ─────────────────────────────────────────────────────────────
@app.route('/cc_clientes')
@login_required
def cc_clientes():
    clientes = db.get_cc_clientes_resumen()
    hoy = date.today().isoformat()
    # Add days_to_vto info
    return render_template('cc_clientes.html', clientes=clientes, hoy=hoy)

@app.route('/cc_clientes/<int:cid>')
@login_required
def cc_cliente_detalle(cid):
    cliente = db.get_cliente(cid)
    movs = db.get_cc_movimientos(cid)
    saldo = db.get_saldo_cliente(cid)
    return render_template('cc_cliente_detalle.html', cliente=cliente, movs=movs, saldo=saldo)

@app.route('/cc_clientes/<int:cid>/mov', methods=['POST'])
@login_required
def cc_cliente_mov_nuevo(cid):
    db.add_cc_mov(cid, request.form.to_dict())
    flash('✅ Movimiento registrado', 'success')
    return redirect(url_for('cc_cliente_detalle', cid=cid))

@app.route('/clientes/nuevo', methods=['POST'])
@login_required
def cliente_nuevo():
    check = db.check_tier_limit('clientes')
    if not check['ok']:
        flash(
            f'⚠ Límite del plan Básico: máximo {check["limite"]} clientes. '
            f'Tenés {check["actual"]}. Actualizá a Pro para agregar más.',
            'warning'
        )
        return redirect(request.referrer or url_for('cc_clientes'))
    db.add_cliente(request.form.to_dict())
    flash('✅ Cliente creado', 'success')
    return redirect(request.referrer or url_for('cc_clientes'))

@app.route('/clientes/<int:cid>/editar', methods=['POST'])
@login_required
def cliente_editar(cid):
    db.update_cliente(cid, request.form.to_dict())
    flash('✅ Cliente actualizado', 'success')
    return redirect(request.referrer or url_for('cc_clientes'))

# ─── CC PROVEEDORES ──────────────────────────────────────────────────────────
@app.route('/cc_proveedores')
@login_required
def cc_proveedores():
    provs = db.get_proveedores()
    facturas = db.get_facturas_proveedores()
    # Summary per proveedor
    total_deuda = sum(max(0, f['importe']-f['pagado']) for f in facturas if f['estado']!='PAGADA')
    vencidas = [f for f in facturas if f['estado']=='VENCIDA']
    por_vencer = [f for f in facturas if f['estado']=='POR VENCER']
    return render_template('cc_proveedores.html', proveedores=provs, facturas=facturas,
                           total_deuda=total_deuda, vencidas=vencidas, por_vencer=por_vencer)

@app.route('/cc_proveedores/factura', methods=['POST'])
@login_required
def factura_nueva():
    db.add_factura_proveedor(request.form.to_dict())
    flash('✅ Factura registrada', 'success')
    return redirect(url_for('cc_proveedores'))

@app.route('/cc_proveedores/factura/<int:fid>/pagar', methods=['POST'])
@login_required
def factura_pagar(fid):
    monto = float(request.form.get('monto', 0))
    db.pagar_factura(fid, monto)
    flash('✅ Pago registrado', 'success')
    return redirect(url_for('cc_proveedores'))

@app.route('/proveedores/nuevo', methods=['POST'])
@login_required
def proveedor_nuevo():
    check = db.check_tier_limit('proveedores')
    if not check['ok']:
        flash(
            f'⚠ Límite del plan Básico: máximo {check["limite"]} proveedores. '
            f'Tenés {check["actual"]}. Actualizá a Pro para agregar más.',
            'warning'
        )
        return redirect(url_for('cc_proveedores'))
    db.add_proveedor(request.form.to_dict())
    flash('✅ Proveedor creado', 'success')
    return redirect(url_for('cc_proveedores'))

# ─── ESTADÍSTICAS ────────────────────────────────────────────────────────────
@app.route('/estadisticas')
@login_required
def estadisticas():
    year = int(request.args.get('year', date.today().year))
    ventas_mes = db.get_ventas_por_mes(year)
    meses_labels = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
    ventas_vals = [ventas_mes.get(m, {}).get('total', 0) for m in range(1, 13)]
    tickets_vals = [ventas_mes.get(m, {}).get('tickets', 0) for m in range(1, 13)]
    semanas = db.get_ventas_por_semana(8)
    medios = db.get_ventas_por_medio_pago(year, date.today().month)
    temporadas = db.get_ventas_por_temporada()
    cats = db.get_ventas_por_categoria()
    return render_template('estadisticas.html',
        year=year,
        meses_labels=json.dumps(meses_labels),
        ventas_vals=json.dumps(ventas_vals),
        tickets_vals=json.dumps(tickets_vals),
        semanas=semanas,
        semanas_labels=json.dumps([s['label'] for s in semanas]),
        semanas_vals=json.dumps([s['total'] for s in semanas]),
        medios=medios,
        medios_labels=json.dumps([m['medio_pago'] for m in medios]),
        medios_vals=json.dumps([m['total'] for m in medios]),
        temporadas=temporadas,
        cats=cats,
        cats_labels=json.dumps([c['categoria'] for c in cats[:8]]),
        cats_vals=json.dumps([c['total'] for c in cats[:8]]),
    )

# ─── ANÁLISIS ────────────────────────────────────────────────────────────────
@app.route('/analisis')
@login_required
def analisis():
    fecha_desde = request.args.get('desde', (date.today()-timedelta(days=30)).isoformat())
    fecha_hasta = request.args.get('hasta', date.today().isoformat())
    top = db.get_top_productos(15, fecha_desde, fecha_hasta)
    bottom = db.get_bottom_productos(10)
    temporadas = db.get_ventas_por_temporada()
    rent = db.get_rentabilidad_mes()
    # Gastos por categoría
    gastos_cat = db.q("""SELECT categoria, ROUND(SUM(monto),2) as total, necesario
        FROM gastos WHERE tipo='Gasto' GROUP BY categoria ORDER BY total DESC""")
    return render_template('analisis.html',
        top=top, bottom=bottom, temporadas=temporadas, rent=rent,
        gastos_cat=gastos_cat, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
        top_labels=json.dumps([t['descripcion'][:20] for t in top]),
        top_vals=json.dumps([t['total_pesos'] for t in top]))

# ─── CONFIGURACIÓN ───────────────────────────────────────────────────────────
@app.route('/config', methods=['GET','POST'])
@admin_required
def config():
    if request.method == 'POST':
        db.set_config(request.form.to_dict())
        flash('✅ Configuración guardada', 'success')
        return redirect(url_for('config'))
    cfg = db.get_config()
    cats = db.get_categorias()
    return render_template('config.html', cfg=cfg, categorias=cats)

@app.route('/config/categoria', methods=['POST'])
@admin_required
def config_categoria():
    if db.get_tier() == 'BASICA':
        flash(
            '⚠ Crear categorías personalizadas requiere el plan Pro. '
            'Podés usar las 23 categorías incluidas en el plan Básico.',
            'warning'
        )
        return redirect(url_for('config'))
    nombre = request.form.get('nombre', '').strip()
    if nombre:
        db.add_categoria(nombre)
        flash('✅ Categoría agregada', 'success')
    return redirect(url_for('config'))

@app.route('/config/categoria/eliminar', methods=['POST'])
@login_required
def config_categoria_eliminar():
    nombre = request.form.get('nombre','').strip()
    if nombre:
        db.delete_categoria(nombre)
        flash(f'🗑 Categoría "{nombre}" eliminada', 'warning')
    return redirect(url_for('config'))

# ─── CLIENTES: EDITAR / ELIMINAR ────────────────────────────────────────────
@app.route('/clientes/<int:cid>/eliminar', methods=['POST'])
@login_required
def cliente_eliminar(cid):
    db.delete_cliente(cid)
    flash('🗑 Cliente desactivado', 'warning')
    return redirect(url_for('cc_clientes'))

@app.route('/clientes/<int:cid>/editar_datos', methods=['POST'])
@login_required
def cliente_editar_datos(cid):
    db.update_cliente(cid, request.form.to_dict())
    flash('✅ Cliente actualizado', 'success')
    return redirect(url_for('cc_cliente_detalle', cid=cid))

@app.route('/cc_clientes/<int:mid>/mov/eliminar', methods=['POST'])
@login_required
def cc_mov_eliminar(mid):
    cid = request.form.get('cliente_id', 0)
    db.delete_cc_mov(mid)
    flash('🗑 Movimiento eliminado', 'warning')
    return redirect(url_for('cc_cliente_detalle', cid=cid))

# ─── PROVEEDORES: EDITAR / ELIMINAR ─────────────────────────────────────────
@app.route('/proveedores/<int:pid>/editar', methods=['POST'])
@login_required
def proveedor_editar(pid):
    db.update_proveedor(pid, request.form.to_dict())
    flash('✅ Proveedor actualizado', 'success')
    return redirect(url_for('cc_proveedores'))

@app.route('/proveedores/<int:pid>/eliminar', methods=['POST'])
@login_required
def proveedor_eliminar(pid):
    db.delete_proveedor(pid)
    flash('🗑 Proveedor desactivado', 'warning')
    return redirect(url_for('cc_proveedores'))

@app.route('/cc_proveedores/factura/<int:fid>/eliminar', methods=['POST'])
@login_required
def factura_eliminar(fid):
    db.delete_factura_proveedor(fid)
    flash('🗑 Factura eliminada', 'warning')
    return redirect(url_for('cc_proveedores'))

# ─── API BÚSQUEDA INSTANTÁNEA POS ───────────────────────────────────────────
@app.route('/api/productos/buscar')
@login_required
def api_productos_buscar():
    term = request.args.get('q', '').strip()
    results = db.buscar_productos_pos(term, 12)
    return jsonify([dict(r) for r in results])

# ─── EXPORTS ─────────────────────────────────────────────────────────────────
@app.route('/productos/exportar/excel')
@admin_required
def exportar_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from flask import send_file

        productos = db.get_catalogo_export()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Catálogo'

        # ─ Styles
        verde_h = PatternFill('solid', fgColor='1a2e1a')
        verde_r = PatternFill('solid', fgColor='e8f5e9')
        font_h = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        font_n = Font(name='Calibri', size=10)
        font_b = Font(name='Calibri', size=10, bold=True)
        aln_c = Alignment(horizontal='center', vertical='center', wrap_text=True)
        aln_r = Alignment(horizontal='right', vertical='center')
        bd = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )

        cfg = db.get_config()
        # ─ Title row
        ws.merge_cells('A1:P1')
        tc = ws['A1']
        tc.value = f"📦 Catálogo de Productos — {cfg.get('nombre_negocio', 'Mi Almacén')}"
        tc.font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
        tc.fill = verde_h
        tc.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        ws.merge_cells('A2:P2')
        ws['A2'].value = f"Exportado: {date.today().strftime('%d/%m/%Y')} · Total productos: {len(productos)}"
        ws['A2'].font = Font(name='Calibri', size=9, italic=True, color='555555')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 16

        # ─ Headers
        headers = ['Cód Interno','Cód Barras','Descripción','Categoría','Unidad',
                   'P/Peso','Costo ($)','Precio Vta ($)','IVA','Margen %',
                   'Stock','Stk Mín','Stk Máx','Proveedor','Estado','Valor Stock ($)']
        widths  = [12, 16, 36, 18, 10, 7, 12, 13, 7, 10, 9, 9, 9, 24, 12, 14]

        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=3, column=ci, value=h)
            cell.font = font_h
            cell.fill = PatternFill('solid', fgColor='2e7d32')
            cell.alignment = aln_c
            cell.border = bd
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[3].height = 22

        # ─ Data rows
        estado_colors = {
            'SIN STOCK': 'ffd7d7',
            'CRITICO':   'ffe4cc',
            'BAJO':      'fff8cc',
            'EXCESO':    'd7eaff',
            'NORMAL':    'f0f8f0',
        }
        for ri, p in enumerate(productos, 4):
            fill_c = PatternFill('solid', fgColor=estado_colors.get(p['estado_stock'], 'FFFFFF'))
            row_data = [
                p['codigo_interno'], p['codigo_barras'] or '',
                p['descripcion'], p['categoria'], p['unidad'],
                'Sí' if p['por_peso'] else 'No',
                p['costo'], p['precio_venta'], p['iva'], p['margen_pct'],
                p['stock_actual'], p['stock_minimo'], p['stock_maximo'],
                p['proveedor_habitual'], p['estado_stock'], p['valor_stock']
            ]
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = font_b if ci == 3 else font_n
                cell.fill = fill_c if ci >= 11 else (verde_r if ri % 2 == 0 else PatternFill())
                cell.border = bd
                if ci in (7, 8, 16):   # money cols
                    cell.number_format = '#,##0.00'
                    cell.alignment = aln_r
                elif ci in (10,):      # percent
                    cell.number_format = '0.00'
                    cell.alignment = aln_r
                elif ci in (11,12,13): # stock
                    cell.number_format = '0.##'
                    cell.alignment = aln_r
                else:
                    cell.alignment = aln_c if ci in (6,9) else Alignment(vertical='center')
            ws.row_dimensions[ri].height = 18

        # ─ Freeze panes & autofilter
        ws.freeze_panes = 'A4'
        ws.auto_filter.ref = f'A3:P{len(productos)+3}'

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        fname = f"catalogo_{date.today().isoformat()}.xlsx"
        return send_file(out, download_name=fname,
                         as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except ImportError:
        flash('❌ Instalá openpyxl: pip install openpyxl', 'danger')
        return redirect(url_for('productos'))

@app.route('/productos/exportar/pdf')
@admin_required
def exportar_pdf():
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from io import BytesIO
        from flask import send_file

        productos = db.get_catalogo_export()
        cfg = db.get_config()

        out = BytesIO()
        doc = SimpleDocTemplate(out, pagesize=landscape(A4),
                                leftMargin=12*mm, rightMargin=12*mm,
                                topMargin=15*mm, bottomMargin=12*mm)

        styles = getSampleStyleSheet()
        verde = colors.HexColor('#1a2e1a')
        verde_m = colors.HexColor('#2e7d32')
        verde_cl = colors.HexColor('#e8f5e9')
        rojo = colors.HexColor('#b71c1c')
        naranja = colors.HexColor('#e65100')
        amarillo = colors.HexColor('#f57f17')
        azul = colors.HexColor('#1565c0')
        gris = colors.HexColor('#f5f5f5')

        title_style = ParagraphStyle('title', fontName='Helvetica-Bold',
                                     fontSize=16, textColor=colors.white,
                                     alignment=TA_CENTER, spaceAfter=0)
        sub_style = ParagraphStyle('sub', fontName='Helvetica',
                                   fontSize=9, textColor=colors.HexColor('#555555'),
                                   alignment=TA_CENTER, spaceAfter=6)
        cell_style = ParagraphStyle('cell', fontName='Helvetica', fontSize=7.5,
                                    leading=10, wordWrap='LTR')
        cell_b_style = ParagraphStyle('cellb', fontName='Helvetica-Bold', fontSize=7.5, leading=10)

        story = []

        # ─ Title
        story.append(Paragraph(
            f"📦 Catálogo de Productos — {cfg.get('nombre_negocio', 'Mi Almacén')}",
            title_style))
        story.append(Spacer(1, 2*mm))
        story.append(Paragraph(
            f"Impreso el {date.today().strftime('%d/%m/%Y')} · {len(productos)} productos activos",
            sub_style))
        story.append(Spacer(1, 4*mm))

        # ─ Table
        headers = ['Código', 'Descripción', 'Categoría', 'Unidad',
                   'Costo', 'Precio Vta', 'Margen%',
                   'Stock', 'Stk Mín', 'Estado']
        col_widths = [22*mm, 70*mm, 35*mm, 18*mm,
                      22*mm, 22*mm, 18*mm,
                      16*mm, 16*mm, 22*mm]

        data = [[Paragraph(f'<b>{h}</b>', cell_b_style) for h in headers]]
        estado_bg = {
            'SIN STOCK': rojo,
            'CRITICO':   naranja,
            'BAJO':      amarillo,
            'EXCESO':    azul,
        }
        for p in productos:
            data.append([
                Paragraph(p['codigo_interno'], cell_style),
                Paragraph(f"<b>{p['descripcion']}</b>", cell_b_style),
                Paragraph(p['categoria'], cell_style),
                Paragraph(p['unidad'] + (' (kg)' if p['por_peso'] else ''), cell_style),
                Paragraph(fmt_ars(p['costo']), cell_style),
                Paragraph(fmt_ars(p['precio_venta']), cell_b_style),
                Paragraph(f"{p['margen_pct']:.1f}%", cell_style),
                Paragraph(str(p['stock_actual']), cell_style),
                Paragraph(str(p['stock_minimo']), cell_style),
                Paragraph(p['estado_stock'], cell_b_style),
            ])

        tbl = Table(data, colWidths=col_widths, repeatRows=1)
        style_cmds = [
            ('BACKGROUND', (0,0), (-1,0), verde_m),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 8),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, gris]),
            ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#cccccc')),
            ('TOPPADDING', (0,0), (-1,-1), 2),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
            ('LEFTPADDING', (0,0), (-1,-1), 3),
            ('RIGHTPADDING', (0,0), (-1,-1), 3),
        ]
        # color by stock state
        for ri, p in enumerate(productos, 1):
            bg = estado_bg.get(p['estado_stock'])
            if bg:
                style_cmds.append(('BACKGROUND', (9, ri), (9, ri), bg))
                style_cmds.append(('TEXTCOLOR', (9, ri), (9, ri), colors.white))

        tbl.setStyle(TableStyle(style_cmds))
        story.append(tbl)

        def on_page(canvas, doc):
            canvas.saveState()
            canvas.setFillColor(verde)
            canvas.rect(0, 0, landscape(A4)[0], 8*mm, fill=1, stroke=0)
            canvas.setFillColor(colors.white)
            canvas.setFont('Helvetica', 7)
            canvas.drawCentredString(landscape(A4)[0]/2, 2.5*mm,
                f"Página {doc.page} · Sistema Almacén · {cfg.get('nombre_negocio', '')}")
            canvas.restoreState()

        doc.build(story, onFirstPage=on_page, onLaterPages=on_page)
        out.seek(0)
        fname = f"catalogo_{date.today().isoformat()}.pdf"
        return send_file(out, download_name=fname,
                         as_attachment=False,
                         mimetype='application/pdf')
    except ImportError:
        flash('❌ Instalá reportlab: pip install reportlab', 'danger')
        return redirect(url_for('productos'))


# ─── RESPALDOS ───────────────────────────────────────────────────────────────
@app.route('/respaldo')
@admin_required
def respaldo():
    cfg = db.get_config()
    backup_dir = get_backup_dir()
    archivos = []
    if os.path.isdir(backup_dir):
        for f in sorted(glob.glob(os.path.join(backup_dir, 'almacen_*.db')), reverse=True):
            stat = os.stat(f)
            archivos.append({
                'nombre': os.path.basename(f),
                'ruta': f,
                'tamanio_kb': round(stat.st_size / 1024, 1),
                'fecha': datetime.fromtimestamp(stat.st_mtime).strftime('%d/%m/%Y %H:%M'),
            })
    return render_template('respaldo.html', archivos=archivos, cfg=cfg,
                           backup_dir=backup_dir,
                           ultimo=cfg.get('backup_ultimo', '—'),
                           intervalo=cfg.get('backup_intervalo_h', '24'),
                           keep=cfg.get('backup_keep', '10'))

@app.route('/respaldo/ahora', methods=['POST'])
@admin_required
def respaldo_ahora():
    ok, msg = hacer_backup(manual=True)
    if ok:
        flash(f'✅ Respaldo creado: {os.path.basename(msg)}', 'success')
    else:
        flash(f'❌ Error al respaldar: {msg}', 'danger')
    return redirect(url_for('respaldo'))

@app.route('/respaldo/config', methods=['POST'])
@admin_required
def respaldo_config():
    data = {
        'backup_intervalo_h': request.form.get('backup_intervalo_h', '24'),
        'backup_keep':        request.form.get('backup_keep', '10'),
        'backup_dir':         request.form.get('backup_dir', '').strip(),
    }
    db.set_config(data)
    iniciar_backup_scheduler()   # restart with new interval
    flash('✅ Configuración de respaldos guardada y programación actualizada.', 'success')
    return redirect(url_for('respaldo'))

@app.route('/respaldo/descargar/<nombre>')
@admin_required
def respaldo_descargar(nombre):
    # Sanitize: only allow our filename pattern
    import re
    if not re.match(r'^almacen_\d{4}-\d{2}-\d{2}_\d{2}-\d{2}\.db$', nombre):
        flash('❌ Archivo inválido.', 'danger')
        return redirect(url_for('respaldo'))
    ruta = os.path.join(get_backup_dir(), nombre)
    if not os.path.exists(ruta):
        flash('❌ Archivo no encontrado.', 'danger')
        return redirect(url_for('respaldo'))
    return send_file(ruta, as_attachment=True, download_name=nombre)

@app.route('/respaldo/eliminar/<nombre>', methods=['POST'])
@admin_required
def respaldo_eliminar(nombre):
    import re
    if not re.match(r'^almacen_\d{4}-\d{2}-\d{2}_\d{2}-\d{2}\.db$', nombre):
        flash('❌ Archivo inválido.', 'danger')
    else:
        ruta = os.path.join(get_backup_dir(), nombre)
        if os.path.exists(ruta):
            os.remove(ruta)
            flash(f'🗑 Respaldo eliminado: {nombre}', 'warning')
    return redirect(url_for('respaldo'))

@app.route('/respaldo/restaurar/<nombre>', methods=['POST'])
@admin_required
def respaldo_restaurar(nombre):
    import re
    if not re.match(r'^almacen_\d{4}-\d{2}-\d{2}_\d{2}-\d{2}\.db$', nombre):
        flash('❌ Archivo inválido.', 'danger')
        return redirect(url_for('respaldo'))
    ruta = os.path.join(get_backup_dir(), nombre)
    if not os.path.exists(ruta):
        flash('❌ Archivo no encontrado.', 'danger')
        return redirect(url_for('respaldo'))
    try:
        # Find DB path
        db_path = db.DB_PATH if hasattr(db, 'DB_PATH') else os.path.join(
            os.path.dirname(os.path.abspath(db.__file__)), 'almacen.db')
        # Backup current DB before restoring
        ts = datetime.now().strftime('%Y-%m-%d_%H-%M')
        prev_bkp = os.path.join(get_backup_dir(), f'almacen_{ts}_antes_restaurar.db')
        if os.path.exists(db_path):
            shutil.copy2(db_path, prev_bkp)
        # Restore
        shutil.copy2(ruta, db_path)
        # Force DB re-init on next request
        db._db_initialized = False
        flash(f'✅ Base de datos restaurada desde «{nombre}». '
              f'Se guardó un respaldo del estado anterior.', 'success')
    except Exception as e:
        flash(f'❌ Error al restaurar: {e}', 'danger')
    return redirect(url_for('respaldo'))

# ─── PRODUCTOS: importar desde OpenFoodFacts ─────────────────────────────────
@app.route('/productos/importar')
@admin_required
def productos_importar():
    from services.openfood_importer import get_stats
    stats_db = get_stats()
    return render_template('importar_productos.html', stats_db=stats_db)


@app.route('/productos/importar/seed', methods=['POST'])
@admin_required
def productos_importar_seed():
    """Importa los 360+ productos del dataset local embebido (sin internet)."""
    try:
        from productos_seed import PRODUCTOS_SEED
        productos = [
            {'barcode': p[0], 'name': p[1], 'brand': p[2],
             'category': p[3], 'unit': p[4], 'por_peso': p[5]}
            for p in PRODUCTOS_SEED
        ]
        stats = db.import_productos_bulk(productos)
        flash(
            f'✅ Dataset local importado: {stats["nuevos"]} nuevos, '
            f'{stats["actualizados"]} actualizados, '
            f'{stats["sin_cambios"]} ya existían, '
            f'{stats["errores"]} errores.',
            'success'
        )
    except Exception as e:
        flash(f'❌ Error: {e}', 'danger')
    return redirect(url_for('productos_importar'))


@app.route('/productos/importar/off', methods=['POST'])
@admin_required
def productos_importar_off():
    from services.openfood_importer import import_products, check_connectivity

    # Verificar límite del tier
    check = db.check_tier_limit('productos_off')
    if not check['ok']:
        flash(
            f'⚠ Límite del plan Básico: máximo {check["limite"]} productos desde OpenFoodFacts. '
            f'Ya tenés {check["actual"]}. Actualizá a Pro para importar más.',
            'warning'
        )
        return redirect(url_for('productos_importar'))

    if not check_connectivity(timeout=5):
        flash('❌ Sin conexión a internet. Usá la importación desde el dataset local.', 'warning')
        return redirect(url_for('productos_importar'))

    try:
        # Si es Básica, limitar a lo que le queda hasta el tope
        tier = db.get_tier()
        if tier == 'BASICA':
            from database import TIER_LIMITS
            tope   = TIER_LIMITS['BASICA']['productos_off']
            actual = check.get('actual', 0) if not check['ok'] else \
                     db.q(
                         "SELECT COUNT(*) FROM productos WHERE activo=1 "
                         "AND codigo_barras != '' "
                         "AND (codigo_barras LIKE '779%' OR codigo_barras LIKE '780%')",
                         fetchone=True
                     )[0]
            limit = max(0, tope - actual)
            if limit == 0:
                flash('⚠ Ya alcanzaste el límite de productos OFF del plan Básico.', 'warning')
                return redirect(url_for('productos_importar'))
        else:
            limit = int(request.form.get('limit', 350))

        stats = import_products(limit=limit)
        flash(
            f'✅ OpenFoodFacts: {stats["inserted"]} nuevos importados, '
            f'{stats["skipped"]} ya existían, '
            f'{stats["pages_fetched"]} páginas procesadas.',
            'success'
        )
    except Exception as e:
        flash(f'❌ Error al conectar con OpenFoodFacts: {e}', 'danger')
    return redirect(url_for('productos_importar'))


@app.route('/productos/importar/update', methods=['POST'])
@admin_required
def productos_importar_update():
    """Actualiza productos existentes y agrega nuevos desde OpenFoodFacts."""
    from services.openfood_importer import update_products, check_connectivity
    if not check_connectivity(timeout=5):
        flash('❌ Sin conexión a internet.', 'warning')
        return redirect(url_for('productos_importar'))
    try:
        max_pages = request.form.get('max_pages')
        max_pages = int(max_pages) if max_pages else None
        stats = update_products(max_pages=max_pages)
        flash(
            f'🔄 Actualización: {stats["inserted"]} nuevos, '
            f'{stats["updated"]} actualizados (nombre/marca), '
            f'{stats["skipped"]} sin cambios.',
            'success'
        )
    except Exception as e:
        flash(f'❌ Error: {e}', 'danger')
    return redirect(url_for('productos_importar'))


@app.route('/api/barcode/<barcode>')
@login_required
def api_barcode_lookup(barcode):
    """Busca un producto en OpenFoodFacts por código de barras."""
    try:
        from services.openfood_importer import check_connectivity
        import urllib.request, json
        if not check_connectivity(timeout=4):
            return jsonify({'ok': False, 'msg': 'Sin conexión a internet'})
        url = f'https://world.openfoodfacts.org/api/v2/product/{barcode}.json'
        req = urllib.request.Request(url, headers={'User-Agent': 'nexaralmacen/1.3.0'})
        with urllib.request.urlopen(req, timeout=8) as r:
            data = json.loads(r.read())
        if data.get('status') != 1:
            return jsonify({'ok': False, 'msg': 'Producto no encontrado en OpenFoodFacts'})
        p = data['product']
        name  = (p.get('product_name_es') or p.get('product_name') or '').strip()
        brand = (p.get('brands') or '').split(',')[0].strip()
        from services.openfood_importer import _map_category
        cat = _map_category(p.get('categories_tags', []))
        if name:
            return jsonify({'ok': True, 'producto': {
                'barcode': barcode, 'name': name, 'brand': brand, 'category': cat
            }})
        return jsonify({'ok': False, 'msg': 'Sin nombre en la base de datos'})
    except Exception as e:
        return jsonify({'ok': False, 'msg': str(e)})

# ─── LISTA NEGRA (BLACKLIST) ─────────────────────────────────────────────────

@app.route('/blacklist')
@admin_required
def blacklist():
    items = db.get_blacklist()
    return render_template('blacklist.html', items=items, app_version=APP_VERSION)


@app.route('/blacklist/agregar', methods=['POST'])
@admin_required
def blacklist_agregar():
    barcode = request.form.get('barcode', '').strip()
    nombre  = request.form.get('nombre_producto', '').strip()
    motivo  = request.form.get('motivo', '').strip()
    if not barcode:
        flash('❌ El código de barras es obligatorio.', 'danger')
        return redirect(url_for('blacklist'))
    added = db.add_to_blacklist(barcode, nombre, motivo)
    if added:
        flash(f'🚫 Barcode {barcode} agregado a la lista negra.', 'success')
    else:
        flash(f'⚠ El barcode {barcode} ya estaba en la lista negra.', 'warning')
    return redirect(url_for('blacklist'))


@app.route('/blacklist/eliminar/<barcode>', methods=['POST'])
@admin_required
def blacklist_eliminar(barcode):
    removed = db.remove_from_blacklist(barcode)
    if removed:
        flash(f'✅ Barcode {barcode} desbloqueado. Se importará normalmente.', 'success')
    else:
        flash(f'⚠ No se encontró el barcode {barcode} en la lista negra.', 'warning')
    return redirect(url_for('blacklist'))


@app.route('/blacklist/desde_producto/<int:pid>', methods=['POST'])
@admin_required
def blacklist_desde_producto(pid):
    """Agrega a la blacklist el barcode de un producto existente y lo desactiva."""
    p = db.get_producto(pid)
    if not p:
        flash('❌ Producto no encontrado.', 'danger')
        return redirect(url_for('productos'))
    barcode = p['codigo_barras']
    if not barcode:
        flash('⚠ Este producto no tiene código de barras para bloquear.', 'warning')
        return redirect(url_for('productos'))
    db.add_to_blacklist(barcode, p['descripcion'], 'Bloqueado manualmente desde catálogo')
    # Desactivar el producto
    db.q("UPDATE productos SET activo=0 WHERE id=?", (pid,))
    flash(f'🚫 Producto desactivado y barcode {barcode} bloqueado. No se reimportará.', 'success')
    return redirect(url_for('productos'))


@app.route('/api/blacklist/check/<barcode>')
@login_required
def blacklist_check(barcode):
    is_bl = db.is_blacklisted(barcode)
    return jsonify({'blacklisted': is_bl, 'barcode': barcode})


# ─── CHANGELOG ───────────────────────────────────────────────────────────────
@app.route('/changelog')
@login_required
def changelog():
    entries = db.get_changelog()
    return render_template('changelog.html', entries=entries, app_version=APP_VERSION)


@app.route('/apagar', methods=['POST'])
@admin_required
def apagar_sistema():
    """
    1. Invalida TODAS las sesiones activas (escribe flag en DB).
    2. Limpia la sesión del usuario actual.
    3. Señaliza al launcher que cierre la ventana.
    4. Detiene Flask.
    """
    try:
        db.set_config({'sessions_invalidated_at': db.datetime.now().isoformat()})
    except Exception:
        pass
    session.clear()

    def _shutdown():
        import time as _t
        _t.sleep(1.0)
        os.kill(os.getpid(), signal.SIGTERM)
    threading.Thread(target=_shutdown, daemon=True).start()
    return render_template('apagado.html')


@app.route('/apagar_rapido', methods=['POST'])
def apagar_rapido():
    """Apagado desde la pantalla de login (sin requerir admin)."""
    try:
        db.set_config({'sessions_invalidated_at': db.datetime.now().isoformat()})
    except Exception:
        pass
    session.clear()

    def _shutdown():
        import time as _t
        _t.sleep(1.0)
        os.kill(os.getpid(), signal.SIGTERM)
    threading.Thread(target=_shutdown, daemon=True).start()
    return render_template('apagado.html')

# ─── ACTUALIZACIÓN DEL SISTEMA ──────────────────────────────────────────────

@app.route('/actualizacion')
@admin_required
def actualizacion():
    if db.get_tier() == 'BASICA':
        flash(
            '⚠ Las actualizaciones del sistema están disponibles solo en el plan Pro.',
            'warning'
        )
        return redirect(url_for('dashboard'))
    return render_template('actualizacion.html', app_version=APP_VERSION)


@app.route('/actualizacion/aplicar', methods=['POST'])
@admin_required
def actualizacion_aplicar():
    """
    Aplica una actualización subida como ZIP.
    Reglas:
      - Extrae solo archivos .py y templates/*.html
      - NUNCA sobrescribe almacen.db
      - Aplica migraciones de DB si el nuevo database.py las incluye
    """
    import zipfile, shutil, tempfile, io

    archivo = request.files.get('archivo_zip')
    if not archivo or not archivo.filename.endswith('.zip'):
        flash('❌ Seleccioná un archivo ZIP válido.', 'danger')
        return redirect(url_for('actualizacion'))

    BASE = os.path.dirname(os.path.abspath(__file__))
    EXCLUIDOS = {'almacen.db', 'almacen.db-wal', 'almacen.db-shm'}

    try:
        zip_bytes = archivo.read()
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
            names = zf.namelist()

            # Detectar prefijo (ej: "almacen/") o raíz directa
            prefix = ''
            for n in names:
                if n.endswith('app.py') or n.endswith('database.py'):
                    prefix = n[:n.rfind('/') + 1] if '/' in n else ''
                    break

            actualizados = []
            omitidos     = []

            for name in names:
                # Quitar prefijo base
                rel = name[len(prefix):] if prefix and name.startswith(prefix) else name
                if not rel or rel.endswith('/'):
                    continue

                # Nunca tocar la base de datos
                filename = os.path.basename(rel)
                if filename in EXCLUIDOS:
                    omitidos.append(rel)
                    continue

                # Solo archivos .py, .html y __init__
                ext = os.path.splitext(rel)[1].lower()
                if ext not in ('.py', '.html', '') and filename != '__init__.py':
                    continue

                dest = os.path.join(BASE, rel.replace('/', os.sep))
                os.makedirs(os.path.dirname(dest), exist_ok=True)

                with zf.open(name) as src, open(dest, 'wb') as dst:
                    dst.write(src.read())
                actualizados.append(rel)

        # Re-ejecutar migraciones de DB (solo agrega columnas/tablas nuevas)
        try:
            import importlib
            db_module = importlib.import_module('database')
            importlib.reload(db_module)
            db_module.init_db()
        except Exception as e:
            flash(f'⚠ Archivos actualizados pero error en migración DB: {e}', 'warning')
            return redirect(url_for('actualizacion'))

        flash(
            f'✅ Actualización aplicada correctamente. '
            f'{len(actualizados)} archivos actualizados. '
            f'Reiniciá el sistema para cargar los cambios.',
            'success'
        )

    except zipfile.BadZipFile:
        flash('❌ El archivo ZIP está dañado o no es válido.', 'danger')
    except Exception as e:
        flash(f'❌ Error al aplicar la actualización: {e}', 'danger')

    return redirect(url_for('actualizacion'))


# ─── CLI COMMANDS ─────────────────────────────────────────────────────────────
import click

@app.cli.command("import-products")
@click.option("--limit", default=350, show_default=True,
              help="Cantidad mínima de productos nuevos a importar.")
@click.option("--timeout", default=15, show_default=True,
              help="Segundos de espera por request HTTP.")
def cli_import_products(limit, timeout):
    """
    Importa productos desde OpenFoodFacts (Argentina).

    Ejemplo:
        flask import-products
        flask import-products --limit 500
    """
    from services.openfood_importer import import_products, check_connectivity
    import logging
    logging.basicConfig(level=logging.INFO, format="[OFI] %(levelname)s %(message)s")

    click.echo(f"Verificando conexión con OpenFoodFacts...")
    if not check_connectivity():
        click.secho(
            "ERROR: Sin conexión a internet. Verificá tu red e intentá nuevamente.",
            fg="red"
        )
        raise SystemExit(1)

    click.echo(f"Iniciando importación — objetivo: {limit} productos...")
    try:
        result = import_products(limit=limit, timeout=timeout)
    except Exception as exc:
        click.secho(f"ERROR durante la importación: {exc}", fg="red")
        raise SystemExit(1)

    click.echo("")
    click.secho("─── Resultado ───────────────────────────────", fg="green")
    click.secho(f"  ✅ Insertados:    {result['inserted']}", fg="green")
    click.secho(f"  ⏭  Ya existían:   {result['skipped']}", fg="yellow")
    click.secho(f"  📄 Páginas:       {result['pages_fetched']}")
    click.secho(f"  ❌ Errores red:   {result['errors']}", fg=("red" if result["errors"] else "white"))
    click.echo("─────────────────────────────────────────────")

    if result["inserted"] >= limit:
        click.secho(f"¡Importación exitosa! {result['inserted']} productos disponibles.", fg="green")
    else:
        click.secho(
            f"Se insertaron {result['inserted']}/{limit} productos "
            "(la API puede no tener más resultados para Argentina).",
            fg="yellow"
        )


@app.cli.command("update-products")
@click.option("--pages", "max_pages", default=None, type=int,
              help="Límite de páginas a procesar. Sin valor → todas las disponibles.")
@click.option("--timeout", default=15, show_default=True,
              help="Segundos de espera por request HTTP.")
def cli_update_products(max_pages, timeout):
    """
    Agrega productos nuevos y actualiza nombre/marca de los existentes.
    NUNCA modifica precios, stock ni categoría.

    Ejemplos:
        flask update-products
        flask update-products --pages 5
    """
    from services.openfood_importer import update_products, check_connectivity
    import logging
    logging.basicConfig(level=logging.INFO, format="[OFI] %(levelname)s %(message)s")

    click.echo("Verificando conexión con OpenFoodFacts...")
    if not check_connectivity():
        click.secho(
            "ERROR: Sin conexión a internet. Verificá tu red e intentá nuevamente.",
            fg="red"
        )
        raise SystemExit(1)

    modo = f"máximo {max_pages} páginas" if max_pages else "todas las páginas disponibles"
    click.echo(f"Iniciando actualización — modo: {modo}...")
    try:
        result = update_products(max_pages=max_pages, timeout=timeout)
    except Exception as exc:
        click.secho(f"ERROR durante la actualización: {exc}", fg="red")
        raise SystemExit(1)

    click.echo("")
    click.secho("─── Resultado ───────────────────────────────", fg="green")
    click.secho(f"  ✅ Nuevos:        {result['inserted']}", fg="green")
    click.secho(f"  🔄 Actualizados:  {result['updated']}", fg="cyan")
    click.secho(f"  ⏭  Sin cambios:   {result['skipped']}", fg="yellow")
    click.secho(f"  📄 Páginas:       {result['pages_fetched']}")
    click.secho(f"  ❌ Errores red:   {result['errors']}", fg=("red" if result["errors"] else "white"))
    click.echo("─────────────────────────────────────────────")
    click.secho("Actualización completada.", fg="green")


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
